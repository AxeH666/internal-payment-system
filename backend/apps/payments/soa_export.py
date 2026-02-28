"""
SOA export utilities for PDF and Excel.

Generates immutable snapshot of batch SOA data for export.
All business rules enforced at view layer; this module is pure I/O.
"""

import io
from datetime import datetime

from django.db.models import Sum

from reportlab.lib import colors
from reportlab.lib.pagesizes import letter
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.platypus import (
    SimpleDocTemplate,
    Paragraph,
    Spacer,
    Table,
    TableStyle,
)
from reportlab.lib.enums import TA_CENTER

from apps.payments.models import PaymentBatch


def _get_batch_export_data(batch_id):
    """Fetch batch with requests and SOA versions for export."""
    batch = (
        PaymentBatch.objects.prefetch_related("requests__soa_versions__uploaded_by")
        .select_related("created_by")
        .get(id=batch_id)
    )
    return batch


def export_batch_soa_pdf(batch_id):
    """
    Generate PDF export of batch SOA (immutable snapshot).
    Returns (bytes, filename).
    """
    batch = _get_batch_export_data(batch_id)

    buffer = io.BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=letter,
        rightMargin=inch,
        leftMargin=inch,
        topMargin=inch,
        bottomMargin=inch,
    )

    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        "CustomTitle",
        parent=styles["Heading1"],
        fontSize=16,
        alignment=TA_CENTER,
        spaceAfter=12,
    )
    heading_style = ParagraphStyle(
        "CustomHeading",
        parent=styles["Heading2"],
        fontSize=12,
        spaceAfter=6,
    )

    story = []

    # Title
    story.append(
        Paragraph(
            f"Statement of Account Export - {batch.title}",
            title_style,
        )
    )
    story.append(Spacer(1, 12))

    # Batch info
    story.append(Paragraph("Batch Information", heading_style))
    batch_data = [
        ["Batch ID", str(batch.id)],
        ["Title", batch.title],
        ["Status", batch.status],
        ["Created", batch.created_at.strftime("%Y-%m-%d %H:%M")],
        [
            "Submitted",
            (
                batch.submitted_at.strftime("%Y-%m-%d %H:%M")
                if batch.submitted_at
                else "—"
            ),
        ],
        [
            "Completed",
            (
                batch.completed_at.strftime("%Y-%m-%d %H:%M")
                if batch.completed_at
                else "—"
            ),
        ],
    ]
    t = Table(batch_data, colWidths=[2 * inch, 4 * inch])
    t.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (0, -1), colors.lightgrey),
                ("TEXTCOLOR", (0, 0), (-1, -1), colors.black),
                ("ALIGN", (0, 0), (0, -1), "RIGHT"),
                ("FONTSIZE", (0, 0), (-1, -1), 10),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
                ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
            ]
        )
    )
    story.append(t)
    story.append(Spacer(1, 20))

    # Requests and SOA
    story.append(Paragraph("Payment Requests & SOA Versions", heading_style))

    total = batch.requests.aggregate(s=Sum("amount"))["s"] or 0

    for req in batch.requests.all().order_by("created_at"):
        story.append(
            Paragraph(
                f"<b>{req.beneficiary_name}</b> - "
                f"{req.amount} {req.currency} ({req.status})",
                styles["Normal"],
            )
        )
        story.append(Spacer(1, 4))

        soas = req.soa_versions.all().order_by("version_number")
        if soas:
            soa_rows = [["Version", "Uploaded At", "Uploaded By"]]
            for soa in soas:
                uploader = (
                    (soa.uploaded_by.display_name or soa.uploaded_by.username)
                    if soa.uploaded_by
                    else "System"
                )
                soa_rows.append(
                    [
                        str(soa.version_number),
                        soa.uploaded_at.strftime("%Y-%m-%d %H:%M"),
                        uploader,
                    ]
                )
            soa_table = Table(soa_rows, colWidths=[1 * inch, 2 * inch, 2 * inch])
            soa_table.setStyle(
                TableStyle(
                    [
                        ("BACKGROUND", (0, 0), (-1, 0), colors.lightgrey),
                        ("FONTSIZE", (0, 0), (-1, -1), 9),
                        ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
                        ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
                    ]
                )
            )
            story.append(soa_table)
        else:
            story.append(
                Paragraph("<i>No SOA documents attached</i>", styles["Normal"])
            )
        story.append(Spacer(1, 12))

    # Batch total
    story.append(Paragraph(f"<b>Batch Total: {total}</b>", styles["Normal"]))
    story.append(Spacer(1, 12))
    story.append(
        Paragraph(
            f"<i>Exported on {datetime.utcnow().strftime('%Y-%m-%d %H:%M UTC')}</i>",
            styles["Normal"],
        )
    )

    doc.build(story)
    buffer.seek(0)

    ts = datetime.utcnow().strftime("%Y%m%d_%H%M")
    filename = f"soa_export_{batch.title.replace(' ', '_')}_{ts}.pdf"
    return buffer.read(), filename


def export_batch_soa_excel(batch_id):
    """
    Generate Excel export of batch SOA (immutable snapshot).
    Returns (bytes, filename).
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Border, Side

    batch = _get_batch_export_data(batch_id)

    wb = Workbook()
    ws = wb.active
    ws.title = "SOA Export"

    header_font = Font(bold=True)
    header_fill = PatternFill(
        start_color="DDDDDD", end_color="DDDDDD", fill_type="solid"
    )
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    row = 1

    # Batch info
    ws.cell(row=row, column=1, value="Batch Information").font = header_font
    row += 1
    ws.cell(row=row, column=1, value="Batch ID")
    ws.cell(row=row, column=2, value=str(batch.id))
    row += 1
    ws.cell(row=row, column=1, value="Title")
    ws.cell(row=row, column=2, value=batch.title)
    row += 1
    ws.cell(row=row, column=1, value="Status")
    ws.cell(row=row, column=2, value=batch.status)
    row += 1
    ws.cell(row=row, column=1, value="Created")
    ws.cell(row=row, column=2, value=batch.created_at.strftime("%Y-%m-%d %H:%M"))
    row += 2

    # Requests table
    ws.cell(row=row, column=1, value="Payment Requests & SOA").font = header_font
    row += 1
    headers = [
        "Beneficiary",
        "Amount",
        "Currency",
        "Purpose",
        "Status",
        "SOA Version",
        "SOA Uploaded At",
        "SOA Uploaded By",
    ]
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=row, column=col, value=h)
        c.font = header_font
        c.fill = header_fill
        c.border = thin_border
    row += 1

    total = 0
    for req in batch.requests.all().order_by("created_at"):
        total += float(req.amount)
        soas = list(
            req.soa_versions.select_related("uploaded_by").order_by("version_number")
        )
        if soas:
            for soa in soas:
                ws.cell(row=row, column=1, value=req.beneficiary_name)
                ws.cell(row=row, column=2, value=float(req.amount))
                ws.cell(row=row, column=3, value=req.currency)
                ws.cell(row=row, column=4, value=req.purpose)
                ws.cell(row=row, column=5, value=req.status)
                ws.cell(row=row, column=6, value=soa.version_number)
                ws.cell(
                    row=row, column=7, value=soa.uploaded_at.strftime("%Y-%m-%d %H:%M")
                )
                uploader = (
                    (soa.uploaded_by.display_name or soa.uploaded_by.username)
                    if soa.uploaded_by
                    else "System"
                )
                ws.cell(row=row, column=8, value=uploader)
                for col in range(1, 9):
                    ws.cell(row=row, column=col).border = thin_border
                row += 1
        else:
            ws.cell(row=row, column=1, value=req.beneficiary_name)
            ws.cell(row=row, column=2, value=float(req.amount))
            ws.cell(row=row, column=3, value=req.currency)
            ws.cell(row=row, column=4, value=req.purpose)
            ws.cell(row=row, column=5, value=req.status)
            ws.cell(row=row, column=6, value="—")
            ws.cell(row=row, column=7, value="—")
            ws.cell(row=row, column=8, value="—")
            for col in range(1, 9):
                ws.cell(row=row, column=col).border = thin_border
            row += 1

    row += 1
    ws.cell(row=row, column=1, value="Batch Total").font = header_font
    ws.cell(row=row, column=2, value=total).font = header_font
    row += 1
    ws.cell(
        row=row,
        column=1,
        value=f"Exported on {datetime.utcnow().strftime('%Y-%m-%d %H:%M UTC')}",
    ).font = Font(italic=True)

    # Column widths
    from openpyxl.utils import get_column_letter

    for col in range(1, 9):
        ws.column_dimensions[get_column_letter(col)].width = 18

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    ts = datetime.utcnow().strftime("%Y%m%d_%H%M")
    filename = f"soa_export_{batch.title.replace(' ', '_')}_{ts}.xlsx"
    return buffer.read(), filename
